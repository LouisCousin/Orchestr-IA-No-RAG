"""Journal consolidé des décisions HITL (Human-In-The-Loop).

Phase 3 : centralise toutes les interventions humaines dans un journal
structuré et exportable.
"""

import logging
from datetime import datetime
from pathlib import Path
from typing import Optional

from src.utils.config import ROOT_DIR
from src.utils.file_utils import load_json, save_json

logger = logging.getLogger("orchestria")

JOURNAL_PATH = ROOT_DIR / "config" / "hitl_journal.json"


class HITLJournal:
    """Journal consolidé des interventions humaines."""

    def __init__(self, journal_path: Optional[Path] = None):
        self.journal_path = journal_path or JOURNAL_PATH
        self._entries: list[dict] = []
        self._next_id: int = 1
        self._load()

    def _load(self) -> None:
        """Charge le journal depuis le disque."""
        if self.journal_path.exists():
            try:
                data = load_json(self.journal_path)
                self._entries = data.get("entries", [])
                self._next_id = data.get("next_id", len(self._entries) + 1)
            except Exception as e:
                logger.warning(f"Erreur chargement journal HITL : {e}")
                self._entries = []
                self._next_id = 1
        else:
            self._entries = []
            self._next_id = 1

    def _save(self) -> None:
        """Sauvegarde le journal sur disque."""
        data = {
            "entries": self._entries,
            "next_id": self._next_id,
        }
        save_json(self.journal_path, data)

    def log_intervention(
        self,
        project_name: str,
        checkpoint_type: str,
        intervention_type: str,
        section_id: Optional[str] = None,
        original_content: Optional[str] = None,
        modified_content: Optional[str] = None,
        delta_summary: str = "",
        model_used: str = "",
        quality_score_before: Optional[float] = None,
        quality_score_after: Optional[float] = None,
    ) -> dict:
        """Ajoute une entrée au journal.

        Args:
            project_name: Nom du projet.
            checkpoint_type: Type de checkpoint (PLAN_VALIDATION, etc.).
            intervention_type: Nature de l'intervention (accept, reject, modify, skip, regenerate, reorder).
            section_id: Identifiant de la section concernée.
            original_content: Contenu avant intervention (extrait ou hash).
            modified_content: Contenu après intervention (extrait ou hash).
            delta_summary: Résumé de la modification.
            model_used: Modèle IA utilisé.
            quality_score_before: Score qualité avant.
            quality_score_after: Score qualité après.

        Returns:
            L'entrée créée.
        """
        entry = {
            "id": self._next_id,
            "project_name": project_name,
            "section_id": section_id,
            "checkpoint_type": checkpoint_type,
            "intervention_type": intervention_type,
            "original_content": (original_content[:500] if original_content else None),
            "modified_content": (modified_content[:500] if modified_content else None),
            "delta_summary": delta_summary,
            "model_used": model_used,
            "quality_score_before": quality_score_before,
            "quality_score_after": quality_score_after,
            "timestamp": datetime.now().isoformat(),
        }
        self._entries.append(entry)
        self._next_id += 1
        self._save()
        logger.info(
            f"HITL Journal : {intervention_type} sur {checkpoint_type}"
            + (f" (section {section_id})" if section_id else "")
        )
        return entry

    def get_interventions(
        self,
        project: Optional[str] = None,
        section: Optional[str] = None,
        intervention_type: Optional[str] = None,
    ) -> list[dict]:
        """Filtre les entrées du journal.

        Args:
            project: Filtrer par nom de projet.
            section: Filtrer par section_id.
            intervention_type: Filtrer par type d'intervention.

        Returns:
            Liste des entrées correspondantes.
        """
        results = self._entries
        if project:
            results = [e for e in results if e.get("project_name") == project]
        if section:
            results = [e for e in results if e.get("section_id") == section]
        if intervention_type:
            results = [e for e in results if e.get("intervention_type") == intervention_type]
        return results

    def get_statistics(self, project: Optional[str] = None) -> dict:
        """Retourne les statistiques agrégées.

        Args:
            project: Filtrer par projet (None = tous).

        Returns:
            Dict avec les statistiques.
        """
        entries = self.get_interventions(project=project)
        if not entries:
            return {
                "total": 0,
                "by_type": {},
                "by_checkpoint": {},
                "modification_rate": 0.0,
            }

        by_type: dict[str, int] = {}
        by_checkpoint: dict[str, int] = {}
        for e in entries:
            itype = e.get("intervention_type", "unknown")
            ctype = e.get("checkpoint_type", "unknown")
            by_type[itype] = by_type.get(itype, 0) + 1
            by_checkpoint[ctype] = by_checkpoint.get(ctype, 0) + 1

        total = len(entries)
        modifications = by_type.get("modify", 0)
        modification_rate = modifications / total if total > 0 else 0.0

        return {
            "total": total,
            "by_type": by_type,
            "by_checkpoint": by_checkpoint,
            "modification_rate": modification_rate,
        }

    def export_to_excel(self, filepath: Path, project: Optional[str] = None) -> Path:
        """Exporte le journal en Excel.

        Args:
            filepath: Chemin du fichier Excel de sortie.
            project: Filtrer par projet (None = tous).

        Returns:
            Chemin du fichier créé.
        """
        from openpyxl import Workbook

        entries = self.get_interventions(project=project)
        if not entries:
            entries = [{"id": "", "project_name": "", "message": "Aucune entrée"}]

        # Grouper par projet
        projects = {}
        for e in entries:
            pname = e.get("project_name") or "Sans projet"
            if pname not in projects:
                projects[pname] = []
            projects[pname].append(e)

        filepath.parent.mkdir(parents=True, exist_ok=True)
        wb = Workbook()
        # Remove the default sheet created by Workbook()
        wb.remove(wb.active)

        for pname, pentries in projects.items():
            sheet_name = pname[:31]  # Excel limits sheet names to 31 chars
            ws = wb.create_sheet(title=sheet_name)
            if pentries:
                headers = list(pentries[0].keys())
                for col_idx, header in enumerate(headers, 1):
                    ws.cell(row=1, column=col_idx, value=header)
                for row_idx, row in enumerate(pentries, 2):
                    for col_idx, header in enumerate(headers, 1):
                        ws.cell(row=row_idx, column=col_idx, value=row.get(header, ""))

        wb.save(str(filepath))

        logger.info(f"Journal HITL exporté : {filepath}")
        return filepath

    @property
    def entries(self) -> list[dict]:
        """Accès en lecture aux entrées."""
        return list(self._entries)
